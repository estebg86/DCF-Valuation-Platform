"""
Excel Valuation Generator
Creates professional Excel valuation models
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

from dcf_model import DCFModel

def generate_excel_valuation(company_data: dict, assumptions: dict) -> str:
    """Generate a complete Excel valuation model"""
    
    ticker = company_data['ticker']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"DCF_{ticker}_{timestamp}.xlsx"
    filepath = os.path.join('./data', filename)
    
    # Initialize DCF model
    dcf = DCFModel(company_data, assumptions)
    
    # Calculate valuation
    valuation_result = dcf.calculate_dcf_valuation()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    create_summary_sheet(wb, company_data, assumptions, valuation_result)
    create_historical_sheet(wb, company_data, dcf.historical_data)
    create_projections_sheet(wb, valuation_result['projections'], assumptions)
    create_dcf_sheet(wb, valuation_result)
    create_sensitivity_sheet(wb, dcf)
    create_assumptions_sheet(wb, assumptions)
    
    # Save workbook
    wb.save(filepath)
    
    return filename

def create_summary_sheet(wb: Workbook, company_data: dict, assumptions: dict, 
                        valuation_result: dict):
    """Create valuation summary sheet"""
    ws = wb.create_sheet("Valuation Summary", 0)
    
    # Styling
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    blue_font = Font(color='0000FF')
    
    # Title
    ws['A1'] = f"{company_data['entity_name']} ({company_data['ticker']})"
    ws['A1'].font = title_font
    
    ws['A2'] = f"Valuation Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A3'] = f"CIK: {company_data['cik']}"
    ws['A4'] = f"Industry: {company_data.get('sic_description', 'N/A')}"
    
    # Valuation Results
    row = 6
    ws[f'A{row}'] = "VALUATION RESULTS"
    ws[f'A{row}'].font = header_font
    
    results = [
        ('Enterprise Value', valuation_result['enterprise_value'], '$'),
        ('Equity Value', valuation_result['equity_value'], '$'),
        ('Shares Outstanding', valuation_result['shares_outstanding'], ''),
        ('Value per Share', valuation_result['value_per_share'], '$'),
        ('Current Price', valuation_result.get('current_price'), '$'),
        ('Upside/(Downside)', valuation_result.get('upside_downside'), '%'),
    ]
    
    row += 1
    for label, value, fmt in results:
        ws[f'A{row}'] = label
        if value is not None:
            if fmt == '$':
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '$#,##0.00'
            elif fmt == '%':
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '0.0%'
            else:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '#,##0'
        row += 1
    
    # Key Assumptions
    row += 2
    ws[f'A{row}'] = "KEY ASSUMPTIONS"
    ws[f'A{row}'].font = header_font
    
    assumptions_display = [
        ('WACC', valuation_result['wacc'], '%'),
        ('Terminal Growth Rate', assumptions.get('terminal_growth_rate', 0.025), '%'),
        ('Tax Rate', assumptions.get('tax_rate', 0.21), '%'),
        ('Projection Years', assumptions.get('projection_years', 10), ''),
    ]
    
    row += 1
    for label, value, fmt in assumptions_display:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = blue_font
        if value is not None:
            if fmt == '%':
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '0.0%'
            else:
                ws[f'B{row}'] = value
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

def create_historical_sheet(wb: Workbook, company_data: dict, historical_df: pd.DataFrame):
    """Create historical financials sheet"""
    ws = wb.create_sheet("Historical Data")
    
    # Title
    ws['A1'] = f"{company_data['ticker']} - Historical Financials"
    ws['A1'].font = Font(bold=True, size=12)
    
    ws['A2'] = "Source: SEC EDGAR (10-K filings)"
    ws['A2'].font = Font(italic=True, size=9)
    
    # Write data
    if not historical_df.empty:
        # Headers
        row = 4
        ws[f'A{row}'] = 'Metric'
        ws[f'A{row}'].font = Font(bold=True)
        
        years = historical_df.index.tolist()
        for col_idx, year in enumerate(years, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = str(year)
            cell.font = Font(bold=True)
        
        # Data rows
        row += 1
        for metric in historical_df.columns:
            ws.cell(row=row, column=1, value=metric)
            
            for col_idx, year in enumerate(years, start=2):
                value = historical_df.loc[year, metric]
                cell = ws.cell(row=row, column=col_idx)
                if pd.notna(value):
                    cell.value = float(value)
                    cell.number_format = '#,##0'
            row += 1
    
    ws.column_dimensions['A'].width = 30

def create_projections_sheet(wb: Workbook, projections_df: pd.DataFrame, 
                            assumptions: dict):
    """Create financial projections sheet"""
    ws = wb.create_sheet("Projections")
    
    # Title
    ws['A1'] = "Financial Projections"
    ws['A1'].font = Font(bold=True, size=12)
    
    # Write projections
    if not projections_df.empty:
        # Headers
        row = 3
        ws[f'A{row}'] = 'Year'
        ws[f'A{row}'].font = Font(bold=True)
        
        metrics = projections_df.columns.tolist()
        for col_idx, metric in enumerate(metrics, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = metric
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)
        
        # Data rows
        row += 1
        for year_idx, year in enumerate(projections_df.index):
            ws.cell(row=row, column=1, value=int(year))
            
            for col_idx, metric in enumerate(metrics, start=2):
                value = projections_df.loc[year, metric]
                cell = ws.cell(row=row, column=col_idx)
                if pd.notna(value):
                    cell.value = float(value)
                    if metric in ['Revenue', 'OperatingIncome', 'FCF', 'PV_FCF', 'CapEx', 
                                 'Depreciation', 'NOPAT', 'Taxes', 'WC_Change']:
                        cell.number_format = '$#,##0'
                    else:
                        cell.number_format = '0.0000'
            row += 1
    
    ws.column_dimensions['A'].width = 12

def create_dcf_sheet(wb: Workbook, valuation_result: dict):
    """Create DCF calculation sheet"""
    ws = wb.create_sheet("DCF Valuation")
    
    # Title
    ws['A1'] = "DCF Valuation Calculation"
    ws['A1'].font = Font(bold=True, size=12)
    
    # Components
    row = 3
    components = [
        ('PV of Projection Period', valuation_result['pv_projection_period']),
        ('Terminal Value', valuation_result['terminal_value']),
        ('PV of Terminal Value', valuation_result['pv_terminal_value']),
        ('Enterprise Value', valuation_result['enterprise_value']),
        ('(+) Cash', 0),  # Would need to be passed in assumptions
        ('(-) Debt', 0),
        ('Equity Value', valuation_result['equity_value']),
        ('Shares Outstanding', valuation_result['shares_outstanding']),
        ('Value per Share', valuation_result['value_per_share']),
    ]
    
    for label, value in components:
        ws[f'A{row}'] = label
        if value is not None:
            ws[f'B{row}'] = value
            if label in ['Shares Outstanding']:
                ws[f'B{row}'].number_format = '#,##0'
            else:
                ws[f'B{row}'].number_format = '$#,##0.00'
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

def create_sensitivity_sheet(wb: Workbook, dcf: DCFModel):
    """Create sensitivity analysis sheet"""
    ws = wb.create_sheet("Sensitivity Analysis")
    
    # Title
    ws['A1'] = "Sensitivity Analysis: Value per Share"
    ws['A1'].font = Font(bold=True, size=12)
    
    # Generate sensitivity table
    sensitivity_df = dcf.sensitivity_analysis()
    
    # Write table
    row = 3
    ws[f'A{row}'] = 'WACC \\ Terminal Growth'
    ws[f'A{row}'].font = Font(bold=True)
    
    # Column headers (terminal growth rates)
    for col_idx, growth_rate in enumerate(sensitivity_df.columns, start=2):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = growth_rate
        cell.number_format = '0.0%'
        cell.font = Font(bold=True)
    
    # Row headers and values
    row += 1
    for wacc in sensitivity_df.index:
        ws.cell(row=row, column=1, value=wacc)
        ws.cell(row=row, column=1).number_format = '0.0%'
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        for col_idx, growth_rate in enumerate(sensitivity_df.columns, start=2):
            value = sensitivity_df.loc[wacc, growth_rate]
            cell = ws.cell(row=row, column=col_idx)
            if pd.notna(value):
                cell.value = float(value)
                cell.number_format = '$#,##0.00'
        row += 1
    
    ws.column_dimensions['A'].width = 15

def create_assumptions_sheet(wb: Workbook, assumptions: dict):
    """Create detailed assumptions sheet"""
    ws = wb.create_sheet("Assumptions")
    
    # Title
    ws['A1'] = "Model Assumptions"
    ws['A1'].font = Font(bold=True, size=12)
    
    ws['A2'] = "Source: User Input / SEC EDGAR"
    ws['A2'].font = Font(italic=True, size=9)
    
    # Write assumptions
    row = 4
    blue_font = Font(color='0000FF')
    
    for key, value in assumptions.items():
        ws[f'A{row}'] = key.replace('_', ' ').title()
        ws[f'A{row}'].font = blue_font
        
        if value is not None:
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = value
                if key in ['terminal_growth_rate', 'tax_rate', 'equity_risk_premium', 
                          'risk_free_rate', 'operating_margin']:
                    ws[f'B{row}'].number_format = '0.0%'
            elif isinstance(value, list):
                ws[f'B{row}'] = str(value)
            else:
                ws[f'B{row}'] = str(value)
        
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
